import openpyxl
from openpyxl.cell.cell import get_column_letter
from openpyxl.styles.colors import Color

##from openpyxl.styles import Font, Alignment

wb = openpyxl.Workbook()
##wb = x.
wb.create_sheet('FILTRO')
ws = wb['FILTRO']

r1 = ['primeiro','segundo','terceiro','quarto','quinto']

# ESTILOS
headFont = openpyxl.styles.Font(name='Consolas', bold=True, color="434343")
headAlign = openpyxl.styles.Alignment(horizontal='center')
headFill = openpyxl.styles.PatternFill('solid', fgColor="f3f3f3")

thin = openpyxl.styles.Side(border_style="thin", color="000000")
headborder = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

# LOOP para cada coluna
for i in range(len(r1)):
    # COORDENADAS
    letter = get_column_letter(i+1)
    cellCoord = letter + str(1)

    ws.column_dimensions[letter].width = 20 # LARGURA dos campos
    
    ws[cellCoord].font = headFont
    ws[cellCoord].alignment = headAlign
    ws[cellCoord].fill = headFill
    ws[cellCoord].border = headborder
    
    
    ws[cellCoord] = r1[i]
    
wb.save(r'.\primeiro_xlsx.xlsx')
